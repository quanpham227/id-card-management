import os
from typing import List, Optional
from datetime import datetime
from datetime import date, timedelta
from fastapi.responses import StreamingResponse # Để xuất file
from fastapi.responses import FileResponse
import os
import tempfile
from fastapi import APIRouter, Depends, HTTPException, status, BackgroundTasks
from sqlalchemy.orm import Session
from sqlalchemy import func, desc
from fastapi.responses import StreamingResponse
import io 
from openpyxl import Workbook # <--- Import thư viện Excel
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill, Alignment #<--- Import các style cho Excel
from typing import Optional
from app.config import TICKET_IMAGES_DIR

from app.database import get_db
from app import models, schemas
from app.oauth2 import get_current_user

router = APIRouter(prefix="/api/tickets", tags=["Tickets"])


# --- 1. TẠO TICKET (Dùng JSON Schema thay vì Form để đồng bộ với ticket_upload.py) ---
@router.post("", response_model=schemas.TicketResponse)
async def create_ticket(
    ticket_in: schemas.TicketCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    file_paths = []
    if ticket_in.attachment_url:
        file_paths = [
            path.strip() for path in ticket_in.attachment_url.split(",") if path.strip()
        ]

    # ĐẢM BẢO các tham số truyền vào đây khớp 100% với Column trong models.py
    new_ticket = models.Ticket(
        title=ticket_in.title,
        description=ticket_in.description,
        category_id=ticket_in.category_id,  # Đã sửa từ 'category' thành 'category_id'
        asset_id=ticket_in.asset_id,  # Đã thêm asset_id
        priority=str(ticket_in.priority),
        status="Open",
        attachments=file_paths,
        requester_id=current_user.id,
        created_at=datetime.now(),
    )

    try:
        db.add(new_ticket)
        db.commit()
        db.refresh(new_ticket)
        return new_ticket
    except Exception as e:
        db.rollback()
        print(f"Database Error Details: {e}")  # Debug lỗi chi tiết ra console
        raise HTTPException(
            status_code=500, detail="Could not save ticket to database."
        )


# --- 2. LẤY DANH SÁCH "MY TICKETS" ---
# app/routers/tickets.py


# Sửa hàm get_my_tickets
@router.get(
    "/my-tickets", response_model=schemas.TicketPaginationResponse
)  # Đổi Schema trả về
def get_my_tickets(
    page: int = 1,
    size: int = 10,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    # Query cơ bản: Lọc theo người tạo (requester_id)
    query = (
        db.query(models.Ticket)
        .options(joinedload(models.Ticket.ticket_category))
        .filter(models.Ticket.requester_id == current_user.id)
    )

    # Đếm tổng số
    total_records = query.count()

    # Phân trang
    skip = (page - 1) * size
    tickets = (
        query.order_by(desc(models.Ticket.created_at)).offset(skip).limit(size).all()
    )

    # Trả về đúng format Schema phân trang
    return {"items": tickets, "total": total_records, "page": page, "size": size}


# --- 3. QUẢN LÝ TICKET (Cho Admin/Manager) ---
@router.get("/manage", response_model=schemas.TicketPaginationResponse)
def manage_tickets(
    status: Optional[str] = None,
    priority: Optional[str] = None,
    page: int = 1,  # Trang hiện tại (Mặc định trang 1)
    size: int = 10,  # Số lượng mỗi trang (Mặc định 10)
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    if current_user.role not in ["Admin", "Manager"]:
        raise HTTPException(status_code=403, detail="Permission denied")

    # 1. Tạo Query cơ bản
    query = db.query(models.Ticket).options(
        joinedload(models.Ticket.requester),
        joinedload(models.Ticket.assignee),
        joinedload(models.Ticket.ticket_category),
    )

    # 2. Áp dụng bộ lọc (Filter)
    if status and status != "All":
        query = query.filter(models.Ticket.status == status)
    if priority and priority != "All":
        query = query.filter(models.Ticket.priority == priority)

    # 3. Đếm tổng số bản ghi (Total) TRƯỚC khi cắt trang
    total_records = query.count()

    # 4. Tính toán Skip (Offset) và lấy dữ liệu
    skip = (page - 1) * size
    tickets = (
        query.order_by(desc(models.Ticket.created_at)).offset(skip).limit(size).all()
    )

    # 5. Trả về đúng format Schema
    return {"items": tickets, "total": total_records, "page": page, "size": size}


# --- 4. CHI TIẾT TICKET & COMMENT ---
from sqlalchemy.orm import joinedload



# --- 6. GỬI COMMENT ---
@router.post("/{ticket_id}/comments", response_model=schemas.TicketCommentResponse)
def create_comment(
    ticket_id: int, 
    comment: schemas.TicketCommentCreate, 
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    ticket = db.query(models.Ticket).filter(models.Ticket.id == ticket_id).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")

    # 1. Tạo Comment
    new_comment = models.TicketComment(
        ticket_id=ticket_id,
        user_id=current_user.id,
        content=comment.content,
        type=comment.type # 'Comment' hoặc 'System'
    )
    db.add(new_comment)
    
    # --- LOGIC TỰ ĐỘNG MỞ LẠI TICKET (AUTO RE-OPEN) ---
    # Nếu người comment KHÔNG PHẢI là IT Admin (tức là User thường)
    # Và Ticket đang ở trạng thái "Resolved" hoặc "Cancelled"
    if current_user.role not in ["Admin", "Manager"] and ticket.status in ["Resolved", "Cancelled"]:
        # Tự động đổi trạng thái về "In Progress" để IT chú ý xử lý tiếp
        ticket.status = "In Progress"
        ticket.updated_at = datetime.now()
        
        # Thêm một dòng log hệ thống để báo hiệu
        system_note = models.TicketComment(
            ticket_id=ticket_id,
            user_id=None, # System
            content=f"Ticket has been Re-opened by user comment.",
            type="System"
        )
        db.add(system_note)

    # Nếu là Admin comment thì giữ nguyên trạng thái (ví dụ Admin vào dặn dò thêm)
    
    db.commit()
    db.refresh(new_comment)
    return new_comment


# --- 7. THỐNG KÊ DASHBOARD ---
@router.get("/stats/summary", response_model=schemas.TicketStats)
def get_ticket_stats(
    # Thêm 2 tham số này để nhận ngày từ Frontend
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db), 
    current_user: models.User = Depends(get_current_user)
):
    if current_user.role not in ["Admin", "Manager"]:
        raise HTTPException(status_code=403, detail="Access denied")

    # --- TẠO QUERY CƠ BẢN (BASE QUERY) ---
    # Thay vì query trực tiếp db.query(models.Ticket), ta tạo 1 query có thể lọc
    base_query = db.query(models.Ticket)

    # --- ÁP DỤNG BỘ LỌC NGÀY (NẾU CÓ) ---
    if start_date:
        # So sánh ngày tạo >= start_date
        base_query = base_query.filter(func.date(models.Ticket.created_at) >= start_date)
    if end_date:
        # So sánh ngày tạo <= end_date
        base_query = base_query.filter(func.date(models.Ticket.created_at) <= end_date)

    # --- ĐẾM DỮ LIỆU (LOGIC CŨ ĐƯỢC GIỮ NGUYÊN, CHỈ THAY NGUỒN QUERY) ---
    
    # Tổng số (dựa trên bộ lọc ngày)
    total = base_query.count()

    # Open (Lọc tiếp status từ base_query)
    open_count = base_query.filter(models.Ticket.status == "Open").count()
    
    # In Progress
    in_progress = base_query.filter(models.Ticket.status == "In Progress").count()
    
    # Resolved
    resolved = base_query.filter(models.Ticket.status == "Resolved").count()

    # Critical (Priority '4' và chưa xong)
    critical = (
        base_query
        .filter(models.Ticket.priority == "4", models.Ticket.status != "Resolved")
        .count()
    )

    return {
        "total": total,
        "open": open_count,
        "in_progress": in_progress,
        "resolved": resolved,
        "critical": critical,
    }


# Hàm phụ trợ: Xóa file tạm sau khi tải xong
def remove_file(path: str):
    try:
        os.remove(path)
    except Exception:
        pass

# ==================================================================
# API XUẤT EXCEL (.xlsx) - BẢN HIGH PERFORMANCE (100k+ rows)
# ==================================================================
@router.get("/export")
def export_tickets(
    background_tasks: BackgroundTasks, # <--- Để xóa file sau khi gửi
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    # 1. Kiểm tra quyền hạn
    if current_user.role not in ["Admin", "Manager"]:
        raise HTTPException(status_code=403, detail="Access denied")

    # 2. Chuẩn bị Query (Chưa tải dữ liệu ngay)
    query = db.query(models.Ticket)
    if start_date:
        query = query.filter(func.date(models.Ticket.created_at) >= start_date)
    if end_date:
        query = query.filter(func.date(models.Ticket.created_at) <= end_date)
    
    # Sắp xếp để dữ liệu ra tuần tự đẹp mắt
    query = query.order_by(models.Ticket.created_at.desc())

    # 3. Tạo file tạm trên ổ cứng (Tránh dùng RAM)
    # delete=False để file tồn tại cho đến khi gửi xong
    tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")

    try:
        # 4. Khởi tạo Workbook chế độ Write-Only (Siêu tiết kiệm RAM)
        wb = Workbook(write_only=True)
        ws = wb.create_sheet("Ticket Report")

        # --- TẠO HEADER CÓ STYLE ---
        # Trong chế độ write_only, ta phải tạo Cell object để style trước khi append
        headers = [
            "Ticket ID", "Title", "Requester", "Assignee", 
            "Category", "Status", "Priority", "Created At (VN)", 
            "Resolved At (VN)", "Resolution Note"
        ]
        
        # Định nghĩa Style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        alignment = Alignment(horizontal="center")

        styled_header_cells = []
        for title in headers:
            cell = WriteOnlyCell(ws, value=title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            styled_header_cells.append(cell)
        
        # Ghi dòng Header vào file
        ws.append(styled_header_cells)

        # 5. Lấy dữ liệu theo từng lô (Batching)
        # yield_per(1000): Chỉ tải 1000 dòng vào RAM mỗi lần -> RAM luôn nhẹ
        ticket_cursor = query.yield_per(1000)

        for t in ticket_cursor:
            # --- XỬ LÝ MÚI GIỜ VIỆT NAM (GMT+7) ---
            created_at_vn = ""
            if t.created_at:
                created_at_vn = (t.created_at + timedelta(hours=7)).strftime("%d/%m/%Y %H:%M")
            
            resolved_at_vn = ""
            if t.resolved_at:
                resolved_at_vn = (t.resolved_at + timedelta(hours=7)).strftime("%d/%m/%Y %H:%M")

            # --- XỬ LÝ AN TOÀN DỮ LIỆU ---
            safe_resolution_note = getattr(t, "resolution_note", "") or ""

            # Ghi dòng dữ liệu
            ws.append([
                t.id,
                t.title,
                t.requester.full_name if t.requester else "Unknown",
                t.assignee.full_name if t.assignee else "Unassigned",
                t.ticket_category.name if t.ticket_category else "General",
                t.status,
                t.priority,
                created_at_vn,
                resolved_at_vn,
                safe_resolution_note
            ])

        # 6. Lưu file và đóng
        wb.save(tmp_file.name)
        tmp_file.close() # Đóng để giải phóng khóa file của HDH

        # 7. Trả về file và dọn dẹp
        # Đăng ký task xóa file sau khi response được gửi đi
        background_tasks.add_task(remove_file, tmp_file.name)

        filename = f"Report_{start_date if start_date else 'All'}_to_{end_date if end_date else 'All'}.xlsx"
        
        # Dùng FileResponse để stream file từ ổ cứng
        return FileResponse(
            path=tmp_file.name,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        # Nếu lỗi giữa chừng, xóa file rác
        remove_file(tmp_file.name)
        raise HTTPException(status_code=500, detail=str(e))


# --- 4. CHI TIẾT TICKET & COMMENT ---
@router.get("/{ticket_id}", response_model=schemas.TicketResponse)
def get_ticket_detail(
    ticket_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    # Sử dụng joinedload để lấy luôn thông tin User và Category trong 1 lần truy vấn
    ticket = (
        db.query(models.Ticket)
        .options(
            joinedload(models.Ticket.requester),
            joinedload(models.Ticket.assignee),
            joinedload(
                models.Ticket.ticket_category
            ),  # Đảm bảo tên này khớp với relationship trong models.py
            joinedload(models.Ticket.comments).joinedload(
                models.TicketComment.user
            ),  # Lấy luôn người comment
        )
        .filter(models.Ticket.id == ticket_id)
        .first()
    )

    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")

    # Kiểm tra quyền xem
    is_manager = current_user.role in ["Admin", "Manager"]
    if ticket.requester_id != current_user.id and not is_manager:
        raise HTTPException(
            status_code=403, detail="Not authorized to view this ticket"
        )

    return ticket


# --- 5. CẬP NHẬT TRẠNG THÁI ---
@router.put("/{ticket_id}", response_model=schemas.TicketResponse)
def update_ticket(
    ticket_id: int,
    ticket_update: schemas.TicketUpdate, # Đổi tên biến cho chuẩn
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    # 1. Tìm Ticket
    ticket = db.query(models.Ticket).filter(models.Ticket.id == ticket_id).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")

    is_manager = current_user.role in ["Admin", "Manager"]

    # 2. Phân quyền (Permission Check)
    if not is_manager:
        # User thường chỉ được sửa ticket của chính mình
        if ticket.requester_id != current_user.id:
            raise HTTPException(status_code=403, detail="Access denied")
        
        # User thường chỉ được phép Hủy (Cancelled) hoặc Mở lại (Open)
        # Không được tự ý chuyển sang Resolved hay In Progress
        if ticket_update.status and ticket_update.status not in ["Cancelled", "Open"]:
             raise HTTPException(
                status_code=403, 
                detail="Users can only Cancel or Re-open their tickets"
            )

    # ==================================================================
    # 3. LOGIC NGHIỆP VỤ CAO CẤP (Chỉ áp dụng cho IT/Admin)
    # ==================================================================
    if is_manager:
        
        # --- A. XỬ LÝ RACE CONDITION (Chống tranh giành) ---
        # Nếu đang cố gán người xử lý (assignee_id được gửi lên)
        if ticket_update.assignee_id is not None:
            # Nếu ticket ĐÃ có người nhận, và người đó KHÁC người đang được gán
            if ticket.assignee_id is not None and ticket.assignee_id != ticket_update.assignee_id:
                # Lấy tên người đang giữ ticket
                current_assignee = db.query(models.User).filter(models.User.id == ticket.assignee_id).first()
                assignee_name = current_assignee.full_name if current_assignee else "someone"
                
                raise HTTPException(
                    status_code=409, # 409 Conflict
                    detail=f"Ticket này vừa bị nhận bởi {assignee_name}. Vui lòng reload lại trang!"
                )

        # --- B. XỬ LÝ WORKFLOW 'IN PROGRESS' ---
        # Nếu chuyển trạng thái sang In Progress
        if ticket_update.status == "In Progress":
            # Nếu chưa có người nhận (cả trong DB lẫn trong request gửi lên)
            if ticket.assignee_id is None and ticket_update.assignee_id is None:
                # Tự động gán cho người đang thao tác (Auto Assign to Me)
                ticket.assignee_id = current_user.id

        # --- C. XỬ LÝ WORKFLOW 'RESOLVED' ---
        # Nếu chuyển trạng thái sang Resolved
        if ticket_update.status == "Resolved":
            # Bắt buộc phải có Resolution Note (Gửi lên mới HOẶC đã có sẵn trong DB)
            if not ticket_update.resolution_note and not ticket.resolution_note:
                raise HTTPException(
                    status_code=400, 
                    detail="Vui lòng nhập 'Resolution Note' (Cách khắc phục) trước khi đóng ticket."
                )
            # Cập nhật thời gian hoàn thành
            ticket.resolved_at = datetime.now()

    # ==================================================================
    # 4. THỰC HIỆN CẬP NHẬT (Dynamic Update)
    # ==================================================================
    
    # Lấy ra các trường cần update (loại bỏ các trường None/Null)
    update_data = ticket_update.model_dump(exclude_unset=True)
    
    for key, value in update_data.items():
        # Xử lý riêng cho assignee_id nếu dùng quy ước -1 (như code cũ của bạn)
        if key == "assignee_id" and value == -1:
             setattr(ticket, key, current_user.id)
        else:
             setattr(ticket, key, value)

    # Cập nhật thời gian sửa đổi
    ticket.updated_at = datetime.now()

    # 5. Ghi Log hệ thống (Optional - Nếu có resolution note mới)
    if ticket_update.resolution_note:
        log = models.TicketComment(
            ticket_id=ticket.id,
            user_id=current_user.id,
            content=f"System Update: Ticket Resolved. Note: {ticket_update.resolution_note}",
            type="System",
        )
        db.add(log)

    try:
        db.commit()
        db.refresh(ticket)
    except Exception as e:
        db.rollback()
        # Handle unique violations or other DB errors
        raise HTTPException(status_code=500, detail=str(e))

    return ticket

@router.delete("/{ticket_id}")
def delete_ticket(ticket_id: int, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    # 1. Check quyền: Chỉ Admin mới được xóa
    if current_user.role not in ['Admin', 'Manager']:
        raise HTTPException(status_code=403, detail="Chỉ Admin mới được xóa ticket")

    # 2. Tìm ticket
    ticket = db.query(models.Ticket).filter(models.Ticket.id == ticket_id).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")

    # 3. Xóa ảnh đính kèm trên ổ cứng
    # [SỬA LỖI TẠI ĐÂY]: Đổi attachment_url -> attachments
    if ticket.attachments: 
        try:
            # Kiểm tra xem attachments là chuỗi (string) hay là List (JSON)
            # Nếu trong DB lưu dạng "path1,path2" (String)
            if isinstance(ticket.attachments, str):
                paths = ticket.attachments.split(',')
            # Nếu trong DB lưu dạng ["path1", "path2"] (JSON/List)
            else:
                paths = ticket.attachments

            for path in paths:
                # path có dạng: "uploads/tickets/abc.jpg"
                # Cần trỏ đúng vào thư mục gốc của project
                full_path = os.path.join(os.getcwd(), path) 
                
                if os.path.exists(full_path):
                    try:
                        os.remove(full_path)
                        print(f"Đã xóa file: {full_path}")
                    except Exception as e:
                        print(f"Lỗi xóa file {full_path}: {e}")
        except Exception as e:
            print(f"Lỗi xử lý xóa ảnh: {e}")

    # 4. Xóa dữ liệu trong DB
    db.delete(ticket)
    db.commit()

    return {"message": "Ticket deleted successfully"}

@router.get("/manage/open-only", response_model=List[schemas.TicketResponse])
def get_open_tickets(
    db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)
):
    # Chỉ Admin/Manager mới được gọi
    if current_user.role not in ["Admin", "Manager"]:
        return []

    # Lấy tất cả ticket có status là "Open" và sắp xếp mới nhất lên đầu
    tickets = (
        db.query(models.Ticket)
        .options(joinedload(models.Ticket.requester))
        .filter(models.Ticket.status == "Open")
        .order_by(desc(models.Ticket.created_at))
        .limit(10)
        .all()
    )

    return tickets
